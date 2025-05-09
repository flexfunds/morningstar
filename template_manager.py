import pandas as pd
from pathlib import Path
import logging
import tempfile
import shutil
from typing import Dict, List, Tuple, Optional, Set
from datetime import datetime
from openpyxl import load_workbook
import os
from config import AppConfig
from models import Series

logger = logging.getLogger(__name__)


class TemplateManager:
    """Manages template processing for different distribution types"""

    def __init__(self, config: AppConfig):
        """
        Initialize the template manager

        Args:
            config: Application configuration
        """
        self.config = config
        self.template_dir = Path(config.template_dir)
        self.output_dir = Path(config.output_dir)

        # Create output directory if it doesn't exist
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def cleanup_output_directory(self):
        """Clean up old files from output directory"""
        if self.output_dir.exists():
            for extension in ['*.xls', '*.xlsx']:
                for file in self.output_dir.glob(extension):
                    file.unlink()
            logger.info("Cleaned up old files from output directory")

    def update_morningstar_template(self, nav_dfs: List[Tuple[str, pd.DataFrame]],
                                    date_str: str) -> Path:
        """
        Update Morningstar template with NAV data

        Args:
            nav_dfs: List of (emitter, dataframe) tuples
            date_str: Date string in format MMDDYYYY

        Returns:
            Path to the updated template file
        """
        # Get template path
        template_path = self.template_dir / "Morningstar Performance Template.xls"
        logger.info(f"Updating Morningstar template: {template_path}")

        # Convert date string to datetime
        try:
            date_obj = datetime.strptime(date_str, '%m%d%Y')
            formatted_date = date_obj.strftime('%m/%d/%Y')
        except ValueError:
            raise ValueError(
                f"Invalid date format: {date_str}. Expected MMDDYYYY")

        # Create temporary file
        with tempfile.NamedTemporaryFile(suffix='.xls', delete=False) as tmp_file:
            shutil.copy2(template_path, tmp_file.name)

            try:
                # Load workbook using xlrd for xls format
                import xlrd
                import xlutils.copy

                rb = xlrd.open_workbook(tmp_file.name, formatting_info=True)
                wb = xlutils.copy.copy(rb)
                sheet = wb.get_sheet(0)

                # Combine all DataFrames
                nav_df = pd.concat(
                    [df for _, df in nav_dfs], ignore_index=True)
                logger.info(
                    f"Processing {len(nav_df)} NAV entries for Morningstar template")

                # Update date in cell F2 (row index 1, column index 5)
                # F2 cell for date (column F is index 5)
                sheet.write(1, 5, formatted_date)

                # Start from row 9 (index 8) as per the requirement
                row_idx = 8

                # Update values:
                # - ISIN in column A (index 0)
                # - Valuation date in column F (index 5)
                # - NAV in column H (index 7)
                for _, row in nav_df.iterrows():
                    isin = row['ISIN']
                    nav_value = row['NAV']

                    # Get the valuation date from the CSV data
                    valuation_date = row['Valuation Period-End Date']
                    row_date_formatted = valuation_date.strftime(
                        '%m/%d/%Y') if pd.notna(valuation_date) else formatted_date

                    # Write ISIN to column A (index 0)
                    sheet.write(row_idx, 0, isin)

                    # Write date to column F (index 5) - NAV daily dividend date
                    sheet.write(row_idx, 5, row_date_formatted)

                    # Write NAV to column H (index 7)
                    sheet.write(row_idx, 7, float(nav_value))

                    row_idx += 1

                # Date-based output filename
                output_date = date_obj.strftime('%m.%d.%Y')
                output_path = self.output_dir / \
                    f'Flexfunds ETPs - NAVs {output_date}.xls'

                # Save workbook
                wb.save(str(output_path))

                logger.info(
                    f"Successfully updated Morningstar template and saved to {output_path}")
                return output_path

            except Exception as e:
                logger.error(f"Error updating Morningstar template: {str(e)}")
                raise
            finally:
                try:
                    os.unlink(tmp_file.name)
                except:
                    pass

    def update_six_template(self, nav_dfs: List[Tuple[str, pd.DataFrame]],
                            date_str: str, series_info: Dict[str, Series]) -> Path:
        """
        Update SIX Financial template with NAV data

        Args:
            nav_dfs: List of (emitter, dataframe) tuples
            date_str: Date string in format MMDDYYYY
            series_info: Dictionary of Series objects keyed by ISIN

        Returns:
            Path to the updated template file
        """
        template_path = self.template_dir / "LAM_SFI_Price -SIX Financial Template.xlsx"
        logger.info(f"Updating SIX template: {template_path}")

        # Create temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            shutil.copy2(template_path, tmp_file.name)

            try:
                # Load workbook
                wb = load_workbook(tmp_file.name, data_only=False)
                sheet = wb.active

                # Store the original row height of row 2 (which will become our header)
                original_height = sheet.row_dimensions[2].height

                # Store original column widths for columns B through G (2 through 7)
                original_widths = {}
                for col in range(2, 8):  # B through G
                    col_letter = sheet.cell(row=1, column=col).column_letter
                    if col_letter in sheet.column_dimensions:
                        original_widths[col] = sheet.column_dimensions[col_letter].width

                # Remove the first row while preserving the yellow header row
                sheet.delete_rows(1, 1)

                # Set the height of the new header row (row 1) to match the original
                sheet.row_dimensions[1].height = original_height

                # Delete all columns after G (including H)
                max_col = sheet.max_column
                if max_col > 8:  # If there are columns after G
                    sheet.delete_cols(9, max_col - 8)

                # Delete column A (first column)
                sheet.delete_cols(1, 1)

                # Restore original column widths (now shifted one column left due to deletion of column A)
                for original_col, width in original_widths.items():
                    new_col = original_col - 1  # Shift column index left by 1
                    col_letter = sheet.cell(
                        row=1, column=new_col).column_letter
                    sheet.column_dimensions[col_letter].width = width

                # Combine all DataFrames
                nav_df = pd.concat(
                    [df for _, df in nav_dfs], ignore_index=True)
                logger.info(
                    f"Processing {len(nav_df)} NAV entries for SIX template")

                # Get the most recent date from the input files
                nav_date = nav_df['Valuation Period-End Date'].max()
                formatted_date = nav_date.strftime('%Y.%m.%d')

                # Update NAV values for each series
                rows_updated = 0
                # Start from row 2 (right after the yellow header)
                current_row = 2

                for _, row in nav_df.iterrows():
                    isin = row['ISIN']
                    series = series_info.get(isin)

                    if series:
                        # Fill in the row data in the correct order:
                        # 1. Security name (using series name)
                        sheet.cell(row=current_row, column=1,
                                   value=str(series.series_name).strip())
                        # 2. ISIN
                        sheet.cell(row=current_row, column=2, value=str(isin))
                        # 3. Valuation Date
                        sheet.cell(row=current_row, column=3,
                                   value=row['Valuation Period-End Date'])
                        # 4. Currency
                        sheet.cell(row=current_row, column=4, value=str(
                            series.currency) if series.currency else "USD")
                        # 5. Nav Price
                        sheet.cell(row=current_row, column=5,
                                   value=float(row['NAV']))
                        # 6. Nav Type (always Structured Products)
                        sheet.cell(row=current_row, column=6,
                                   value="Structured Products")
                        # 7. Valuation Frequency
                        sheet.cell(row=current_row, column=7,
                                   value=str(series.nav_frequency.value) if series.nav_frequency else "Daily")

                        current_row += 1
                        rows_updated += 1
                    else:
                        logger.warning(
                            f"No series information found for ISIN {isin}")

                logger.info(f"Updated {rows_updated} rows in the SIX template")

                # Use date-based filename
                output_path = self.output_dir / \
                    f'LAM_SFI_Price - {formatted_date}.xlsx'

                # Save workbook
                wb.save(str(output_path))
                wb.close()

                logger.info(
                    f"Successfully updated SIX template and saved to {output_path}")
                return output_path

            except Exception as e:
                logger.error(f"Error updating SIX template: {str(e)}")
                raise
            finally:
                try:
                    os.unlink(tmp_file.name)
                except:
                    pass
