import logging
from pathlib import Path
import pandas as pd
from typing import List, Tuple, Dict
import xlrd
import xlutils.copy
from openpyxl import load_workbook
import tempfile
import shutil
from datetime import datetime

logger = logging.getLogger(__name__)


class TemplateProcessor:
    def __init__(self, template_dir: Path, output_dir: Path):
        """
        Initialize template processor

        Args:
            template_dir: Directory containing template files
            output_dir: Directory for output files
        """
        self.template_dir = template_dir
        self.output_dir = output_dir

    def _cleanup_output_directory(self):
        """Clean up old files from output directory"""
        if self.output_dir.exists():
            for extension in ['*.xls', '*.xlsx']:
                for file in self.output_dir.glob(extension):
                    file.unlink()
            logger.info("Cleaned up old files from output directory")

    def update_morningstar_template(self, nav_dfs: List[Tuple[str, pd.DataFrame]], date_str: str) -> Path:
        """
        Update Morningstar template with NAV data

        Args:
            nav_dfs: List of (emitter, dataframe) tuples containing NAV data
            date_str: Date string in format MMDDYYYY

        Returns:
            Path: Path to the updated template file
        """
        # Read template
        wb = xlrd.open_workbook(
            self.template_dir / "Morningstar Performance Template.xls",
            formatting_info=True
        )
        template_sheet = wb.sheet_by_name('NAVs')
        wb_output = xlutils.copy.copy(wb)
        sheet_output = wb_output.get_sheet('NAVs')

        # Get column mappings
        header_row = 7
        col_indices = {
            template_sheet.cell_value(header_row, col_idx): col_idx
            for col_idx in range(template_sheet.ncols)
        }

        # Update template with NAV data
        mapping = {
            'Unique Identifier': 'ISIN',
            'NAV/Daily dividend Date': 'Valuation Period-End Date',
            'NAV': 'NAV'
        }

        # Combine all DataFrames
        nav_df = pd.concat([df for _, df in nav_dfs], ignore_index=True)

        # Update the data starting from row 8
        for i, row in nav_df.iterrows():
            row_idx = i + 8
            for template_col, nav_col in mapping.items():
                if template_col in col_indices:
                    col_idx = col_indices[template_col]
                    value = row[nav_col]

                    # Special handling for dates
                    if template_col == 'NAV/Daily dividend Date':
                        if isinstance(value, pd.Timestamp):
                            value = value.strftime('%m/%d/%Y')
                        else:
                            try:
                                date_obj = pd.to_datetime(value)
                                value = date_obj.strftime('%m/%d/%Y')
                            except:
                                pass
                    sheet_output.write(row_idx, col_idx, value)

        # Save updated template
        date_obj = pd.to_datetime(date_str, format='%m%d%Y')
        formatted_date = date_obj.strftime('%m.%d.%Y')
        output_path = self.output_dir / \
            f'Flexfunds ETPs - NAVs {formatted_date}.xls'
        wb_output.save(str(output_path))

        logger.info(
            f"Successfully updated Morningstar template: {output_path}")
        return output_path

    def update_six_template(self, nav_dfs: List[Tuple[str, pd.DataFrame]], date_str: str,
                            series_info: Dict[str, object]) -> Path:
        """
        Update SIX template with NAV data

        Args:
            nav_dfs: List of (emitter, dataframe) tuples containing NAV data
            date_str: Date string in format MMDDYYYY
            series_info: Dictionary mapping ISINs to series information

        Returns:
            Path: Path to the updated template file
        """
        template_path = self.template_dir / "LAM_SFI_Price -SIX Financial Template.xlsx"
        logger.info(
            f"Starting SIX template update using template: {template_path}")

        # Create temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            shutil.copy2(template_path, tmp_file.name)

            try:
                # Load workbook
                wb = load_workbook(tmp_file.name, data_only=False)
                sheet = wb.active

                # Store original formatting
                original_height = sheet.row_dimensions[2].height
                original_widths = {
                    col: sheet.column_dimensions[sheet.cell(
                        row=1, column=col).column_letter].width
                    for col in range(2, 8)
                }

                # Clean up template
                sheet.delete_rows(1, 1)
                sheet.row_dimensions[1].height = original_height

                # Delete unnecessary columns
                max_col = sheet.max_column
                if max_col > 8:
                    sheet.delete_cols(9, max_col - 8)
                sheet.delete_cols(1, 1)

                # Restore column widths
                for original_col, width in original_widths.items():
                    new_col = original_col - 1
                    col_letter = sheet.cell(
                        row=1, column=new_col).column_letter
                    sheet.column_dimensions[col_letter].width = width

                # Combine NAV data
                nav_df = pd.concat(
                    [df for _, df in nav_dfs], ignore_index=True)
                nav_date = nav_df['Valuation Period-End Date'].max()
                formatted_date = nav_date.strftime('%Y.%m.%d')

                # Update NAV values
                current_row = 2
                rows_updated = 0

                for _, row in nav_df.iterrows():
                    isin = row['ISIN']
                    series = series_info.get(isin)

                    if series:
                        # Fill row data
                        sheet.cell(row=current_row, column=1,
                                   value=str(series.series_name).strip())
                        sheet.cell(row=current_row, column=2, value=str(isin))
                        sheet.cell(row=current_row, column=3,
                                   value=row['Valuation Period-End Date'])
                        sheet.cell(row=current_row, column=4, value=str(
                            series.currency) if series.currency else "USD")
                        sheet.cell(row=current_row, column=5,
                                   value=float(row['NAV']))
                        sheet.cell(row=current_row, column=6,
                                   value="Structured Products")
                        sheet.cell(row=current_row, column=7, value=str(
                            series.nav_frequency.value) if series.nav_frequency else "Daily")
                        current_row += 1
                        rows_updated += 1
                    else:
                        logger.warning(
                            f"No series information found for ISIN {isin}")

                logger.info(f"Updated {rows_updated} rows in the SIX template")

                # Save output
                output_path = self.output_dir / \
                    f'LAM_SFI_Price - {formatted_date}.xlsx'
                wb.save(str(output_path))
                wb.close()

                logger.info(
                    f"Successfully updated SIX template: {output_path}")
                return output_path

            except Exception as e:
                logger.error(f"Error updating SIX template: {str(e)}")
                raise
            finally:
                try:
                    os.unlink(tmp_file.name)
                except:
                    logger.warning("Failed to clean up temporary file")
