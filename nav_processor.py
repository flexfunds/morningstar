import pandas as pd
from ftplib import FTP, FTP_TLS
from io import StringIO
import os
from typing import List, Dict, Tuple, Optional, Set, Union
import logging
from pathlib import Path
import xlrd
import xlutils.copy
import ssl
from email_sender import EmailSender
from google_drive_service import GoogleDriveService
import shutil
import tempfile
from contextlib import contextmanager
from db_service import DatabaseService
from datetime import datetime
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from queue import Queue
import threading
import time
from models import Series, SeriesStatus
from ftp_service import FTPService
from template_processor import TemplateProcessor
from sqlalchemy import func

logger = logging.getLogger(__name__)


class NAVProcessor:
    def __init__(self, mode: str = "local", ftp_configs: Dict[str, Dict] = None,
                 smtp_config: Dict = None, drive_config: Dict = None,
                 db_connection_string: str = 'sqlite:///nav_data.db',
                 max_workers: int = 5):
        """
        Initialize the NAV processor

        Args:
            mode (str): Operation mode - either "local" or "remote"
            ftp_configs (Dict[str, Dict]): Dictionary of FTP configurations for each emitter
            smtp_config (Dict): SMTP configuration for email sending
            drive_config (Dict): Google Drive configuration for file syncing
            db_connection_string (str): Database connection string
            max_workers (int): Maximum number of concurrent workers for file operations
        """
        self.mode = mode.lower()
        self.ftp_service = FTPService(ftp_configs) if ftp_configs else None
        self.email_sender = EmailSender(smtp_config) if smtp_config else None
        self.drive_service = GoogleDriveService(
            drive_config['credentials_path']) if drive_config else None
        self.drive_config = drive_config
        self.db_service = DatabaseService(db_connection_string)
        self.max_workers = max_workers

        # Thread-safe queue for Google Drive uploads
        self.upload_queue = Queue()

        # Configure logging
        logging.basicConfig(level=logging.INFO)

        # Set more restrictive logging for noisy libraries
        logging.getLogger('ftplib').setLevel(logging.WARNING)
        logging.getLogger('googleapiclient').setLevel(logging.WARNING)
        logging.getLogger('google_auth_httplib2').setLevel(logging.WARNING)
        logging.getLogger('googleapiclient.discovery').setLevel(
            logging.WARNING)
        logging.getLogger('urllib3').setLevel(logging.WARNING)

        self.logger = logging.getLogger(__name__)

        # Define file structure with existing paths
        self.input_dir = Path("input")
        self.output_dir = Path("output")
        self.template_dir = self.input_dir / "template"
        self.temp_dir = Path(tempfile.gettempdir()) / "nav_processor"

        # Initialize template processor
        self.template_processor = TemplateProcessor(
            self.template_dir, self.output_dir)

        # Create directories if in local mode
        if self.mode == "local":
            self._create_directories()

    def _create_directories(self):
        """Create necessary directories for local mode"""
        for directory in [self.input_dir, self.output_dir, self.temp_dir]:
            directory.mkdir(parents=True, exist_ok=True)

        # Create emitter subdirectories
        if self.ftp_service:
            for emitter in self.ftp_service.config.keys():
                (self.input_dir / emitter).mkdir(exist_ok=True)

    @contextmanager
    def _temp_file_handler(self, filename: str):
        """Context manager for handling temporary files"""
        # Create temp directory if it doesn't exist
        self.temp_dir.mkdir(parents=True, exist_ok=True)

        temp_path = self.temp_dir / filename
        try:
            yield temp_path
        finally:
            if temp_path.exists():
                temp_path.unlink()

    def _read_csv_local(self, filename: str) -> pd.DataFrame:
        """Read CSV file from local directory"""
        file_path = self.input_dir / filename
        return pd.read_csv(file_path)

    def _cleanup_emitter_directory(self, emitter: str):
        """Clean up old files from emitter directory"""
        emitter_dir = self.input_dir / emitter
        if emitter_dir.exists():
            for file in emitter_dir.glob('*.csv'):
                file.unlink()
            self.logger.info(f"Cleaned up old files from {emitter} directory")

    def _read_csv_remote(self, filename: str, emitter: str, temp_file: Path) -> Optional[pd.DataFrame]:
        """Read CSV file from FTP server with improved error handling"""
        ftp_config = self.ftp_service.config.get(emitter)
        if not ftp_config:
            raise ValueError(
                f"No FTP configuration found for emitter {emitter}")

        # Create FTP_TLS instance with custom context
        context = ssl.create_default_context()
        context.check_hostname = False
        context.verify_mode = ssl.CERT_NONE
        context.options |= ssl.OP_NO_TICKET

        with FTP_TLS(context=context) as ftp:
            ftp.encoding = 'utf-8'
            try:
                # Connect and authenticate
                ftp.connect(host=ftp_config['host'], port=21)
                ftp.auth()
                ftp.login(ftp_config['user'], ftp_config['password'])

                # Enable TLS for data channel
                ftp.prot_p()
                ftp.set_pasv(True)

                # Change to directory if specified
                if ftp_config.get('directory'):
                    ftp.cwd(ftp_config['directory'])

                # Download the file
                with open(temp_file, 'wb') as f:
                    ftp.retrbinary(f'RETR {filename}', f.write)

                # Read the CSV file
                try:
                    return pd.read_csv(temp_file)
                except UnicodeDecodeError:
                    return pd.read_csv(temp_file, encoding='latin-1')

            except Exception as e:
                if "550" in str(e):  # File not found
                    return None
                raise

    def cleanup(self):
        """Clean up temporary files"""
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)
            self.temp_dir.mkdir(parents=True, exist_ok=True)
            self.logger.info("Temporary files cleaned up")

    def _cleanup_output_directory(self):
        """Clean up old files from output directory"""
        if self.output_dir.exists():
            for extension in ['*.xls', '*.xlsx']:
                for file in self.output_dir.glob(extension):
                    file.unlink()
            self.logger.info("Cleaned up old files from output directory")

    def _get_input_files(self, date_str: str) -> List[Tuple[str, str]]:
        """Get list of input files with their emitters for a given date."""
        return [
            ('ETPCAP2', f'CAS_Flexfunds_NAV_{date_str} ETPCAP2.csv'),
            ('HFMX', f'CAS_Flexfunds_NAV_{date_str} HFMX.csv'),
            ('IACAP', f'CAS_Flexfunds_NAV_{date_str} IACAP.csv'),
            ('CIX', f'CAS_Flexfunds_NAV_{date_str} CIX.csv'),
            ('DCXPD', f'CAS_Flexfunds_NAV_{date_str} DCXPD.csv'),
            ('ETPCAP2',
             f'CAS_Flexfunds_NAV_{date_str} Wrappers Hybrid ETPCAP2.csv'),
            ('HFMX', f'CAS_Flexfunds_NAV_{date_str} Wrappers Hybrid HFMX.csv'),
            ('IACAP',
             f'CAS_Flexfunds_NAV_{date_str} Wrappers Hybrid IACAP.csv'),
            ('CIX', f'CAS_Flexfunds_NAV_{date_str} Wrappers Hybrid CIX.csv'),
            ('DCXPD',
             f'CAS_Flexfunds_NAV_{date_str} Wrappers Hybrid DCXPD.csv'),
            ('ETPCAP2', f'CAS_Flexfunds_NAV_{date_str} Loan ETPCAP2.csv'),
            ('HFMX', f'CAS_Flexfunds_NAV_{date_str} Loan HFMX.csv'),
            ('IACAP', f'CAS_Flexfunds_NAV_{date_str} Loan IACAP.csv'),
            ('CIX', f'CAS_Flexfunds_NAV_{date_str} Loan CIX.csv'),
            ('DCXPD', f'CAS_Flexfunds_NAV_{date_str} Loan DCXPD.csv')
        ]

    def _process_ftp_file(self, emitter: str, filename: str) -> Optional[pd.DataFrame]:
        """Process a single FTP file download"""
        try:
            # Create a unique temporary file for each download
            temp_file = self.temp_dir / f"{emitter}_{filename}"
            df = self.ftp_service.download_file(emitter, filename, temp_file)

            if df is not None:
                # Clean up the DataFrame
                df = self._clean_dataframe(df)

                # Save to input directory
                input_path = self.input_dir / emitter / filename
                input_path.parent.mkdir(exist_ok=True)
                df.to_csv(input_path, index=False)

                if self.drive_service:
                    # Add to upload queue
                    self.upload_queue.put((emitter, filename))

                # Only log non-empty dataframes
                if not df.empty:
                    self.logger.debug(f"Processed {emitter} file: {filename}")
                return df if not df.empty else None
        except Exception as e:
            if "550" not in str(e):  # Only log non-404 errors
                self.logger.error(f"Error: {filename} from {emitter}")
            return None
        finally:
            # Clean up temp file
            if temp_file.exists():
                try:
                    temp_file.unlink()
                except:
                    pass

    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and standardize DataFrame"""
        # Remove unnamed columns
        unnamed_cols = [col for col in df.columns if 'Unnamed' in col]
        if unnamed_cols:
            df = df.drop(columns=unnamed_cols)

        # Clean up column names
        df.columns = df.columns.str.strip()

        # Convert date column to datetime
        if 'Valuation Period-End Date' in df.columns:
            df['Valuation Period-End Date'] = pd.to_datetime(
                df['Valuation Period-End Date'])

        # Clean up ISIN values
        if 'ISIN' in df.columns:
            df['ISIN'] = df['ISIN'].str.strip()

        # Clean up NAV values
        if 'NAV' in df.columns:
            df['NAV'] = pd.to_numeric(df['NAV'].astype(
                str).str.replace(',', ''), errors='coerce')

        # Standardize frequency values to uppercase
        if 'Frequency' in df.columns:
            df['Frequency'] = df['Frequency'].str.upper()

        return df

    def _upload_worker(self):
        """Worker function for Google Drive uploads with retry logic"""
        while True:
            try:
                emitter, filename = self.upload_queue.get()
                if emitter == "STOP":
                    break

                input_path = self.input_dir / emitter / filename
                if input_path.exists() and self.drive_config.get('input_folder_id'):
                    max_retries = 3
                    retry_count = 0
                    while retry_count < max_retries:
                        try:
                            self.drive_service.upload_file(
                                input_path,
                                self.drive_config['input_folder_id']
                            )
                            # Reduced log verbosity
                            if retry_count > 0:  # Only log if we had to retry
                                self.logger.info(
                                    f"Uploaded {filename} after {retry_count} retries")
                            break
                        except Exception as e:
                            retry_count += 1
                            if retry_count == max_retries:
                                self.logger.error(
                                    f"Failed to upload {filename} after {max_retries} attempts")
                            else:
                                time.sleep(1)  # Wait before retry
            finally:
                self.upload_queue.task_done()

    def _get_series_info(self, isins: Set[str]) -> Dict[str, object]:
        """Get series information from database"""
        with self.db_service.SessionMaker() as session:
            series_info = session.query(Series).filter(
                Series.isin.in_(isins)).all()
            return {s.isin: s for s in series_info}

    def _process_nav_files(self, date_str: str, target_isins: Optional[Set[str]] = None,
                           exclude_isins: Optional[Set[str]] = None) -> List[Tuple[str, pd.DataFrame]]:
        """Process NAV files concurrently and return list of (emitter, dataframe) tuples."""
        nav_dfs = []
        missing_files = []
        input_files = self._get_input_files(date_str)

        # Start upload worker thread
        upload_thread = None
        if self.drive_service:
            upload_thread = threading.Thread(
                target=self._upload_worker, daemon=True)
            upload_thread.start()

        # Process files concurrently with a smaller number of workers
        with ThreadPoolExecutor(max_workers=min(3, self.max_workers)) as executor:
            future_to_file = {
                executor.submit(self._process_ftp_file, emitter, filename): (emitter, filename)
                for emitter, filename in input_files
            }

            for future in as_completed(future_to_file):
                emitter, filename = future_to_file[future]
                try:
                    df = future.result()
                    if df is None:
                        missing_files.append(filename)
                        continue

                    # Clean up the DataFrame
                    df = self._clean_dataframe(df)

                    # Drop rows with missing required values
                    required_cols = ['ISIN', 'NAV',
                                     'Valuation Period-End Date']
                    df = df.dropna(subset=required_cols)

                    # Apply ISIN filters
                    if target_isins:
                        df = df[df['ISIN'].isin(target_isins)]
                    if exclude_isins:
                        df = df[~df['ISIN'].isin(exclude_isins)]

                    if not df.empty:
                        # Add a file_type column to track the file source
                        if 'Wrappers Hybrid' in filename:
                            df['file_type'] = 'hybrid'
                            nav_dfs.append((emitter, df))
                        elif 'Loan' in filename:
                            df['file_type'] = 'loan'
                            nav_dfs.append((emitter, df))
                        else:
                            df['file_type'] = 'standard'
                            nav_dfs.append((emitter, df))

                        self.logger.info(
                            f"Processed {emitter} file: {filename}")

                except Exception as e:
                    self.logger.error(f"Error processing {filename}: {str(e)}")

        # Stop upload worker and wait for remaining uploads
        if upload_thread:
            self.upload_queue.put(("STOP", None))
            self.upload_queue.join()

        if missing_files:
            self.logger.warning(f"Some files were not found: {missing_files}")

        return nav_dfs

    def _get_email_template(self, distribution_type: str, nav_dfs: List[Tuple[str, pd.DataFrame]], output_path: Path) -> Tuple[str, str]:
        """Get email template based on distribution type."""
        if distribution_type.lower() == 'six':
            # Count series by emitter
            emitter_isins = {}  # Dictionary to store sets of ISINs per emitter
            for emitter, df in nav_dfs:
                if emitter not in emitter_isins:
                    emitter_isins[emitter] = set()
                # Add ISINs to the set for this emitter
                emitter_isins[emitter].update(df['ISIN'].unique())
                self.logger.debug(
                    f"Found {len(emitter_isins[emitter])} total unique ISINs for {emitter}")

            # Build the body with emitter counts
            body_lines = [
                "Hi team,",
                "",  # Empty line
                "I hope you are well.",
                "",  # Empty line
                "Attached please find the pricing distribution information for the issuers we function as the calculation agent.",
                ""  # Empty line before counts
            ]

            # Add emitter counts in specific order
            emitters_order = ['IACAP', 'ETPCAP2', 'HFMX', 'CIX', 'DCXPD']
            total_series = 0
            for emitter in emitters_order:
                count = len(emitter_isins.get(emitter, set()))
                total_series += count
                body_lines.append(f"{emitter}: {count}")

            body_lines.extend([
                "",  # Empty line
                f"Total series: {total_series}",
                "",  # Empty line
                "Many thanks,"
            ])

            return ("Pricing distribution - IA Capital, ETPCAP2, HFMX, CIX, and DCXPD",
                    "\n".join(body_lines))
        else:
            # Morningstar template
            return ("FlexFunds Calculation Agent ETPs - Morningstar NAV Update",
                    """Hi team,

I hope you are well.

Please find attached the updated NAV for the pricing distribution process of the different notes for which we function as the calculation agent.

Many thanks,""")

    def _send_email_report(self, output_path: Path, to_emails: List[str], distribution_type: str = 'morningstar', nav_dfs: List[Tuple[str, pd.DataFrame]] = None) -> bool:
        """Send email report with NAV data."""
        if not self.email_sender:
            self.logger.warning("Email sender not configured")
            return False

        # Get email template based on distribution type
        subject, body = self._get_email_template(
            distribution_type, nav_dfs, output_path)

        try:
            email_sent = self.email_sender.send_report(
                to_emails=to_emails,
                subject=subject,
                body=body,
                attachment_path=output_path
            )
            if email_sent:
                self.logger.info("NAV report sent via email successfully")
            else:
                self.logger.error("Failed to send NAV report via email")
            return email_sent
        except Exception as e:
            self.logger.error(f"Error sending email: {str(e)}")
            return False

    def _save_to_database(self, nav_dfs: List[Tuple[str, pd.DataFrame]], distribution_type: str) -> Tuple[int, int, int]:
        """Save NAV data to database and return (added_count, duplicates_count, invalids_count)."""
        total_added = 0
        total_duplicates = 0
        total_invalids = 0

        for emitter, df in nav_dfs:
            if not df.empty:
                # More concise log
                self.logger.info(
                    f"Processing {emitter} data: {len(df)} entries")

                added, duplicates, invalids = self.db_service.save_nav_entries(
                    df, distribution_type, emitter)
                total_added += added
                total_duplicates += duplicates
                total_invalids += invalids

                # Simplify this log to be less cluttered
                if added > 0:
                    self.logger.info(f"{emitter}: Added {added} entries")

        # Log final summary in a more concise format
        if total_added > 0 or total_duplicates > 0:
            self.logger.info(
                f"DB Import: {total_added} added, {total_duplicates} duplicates, {total_invalids} invalid")

        return total_added, total_duplicates, total_invalids

    def _update_six_template(self, nav_dfs: List[Tuple[str, pd.DataFrame]], date_str: str) -> Path:
        """Update SIX Financial template with NAV data and return output path."""
        template_path = self.template_dir / "LAM_SFI_Price -SIX Financial Template.xlsx"
        self.logger.info(
            f"Starting SIX template update using template: {template_path}")

        # Create temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            self.logger.debug(f"Created temporary file: {tmp_file.name}")
            shutil.copy2(template_path, tmp_file.name)

            try:
                # Load workbook
                self.logger.debug("Loading workbook...")
                wb = load_workbook(tmp_file.name, data_only=False)
                sheet = wb.active

                # Store the original row height of row 2 (which will become our header)
                original_height = sheet.row_dimensions[2].height

                # Store original column widths for columns B through G (2 through 7)
                original_widths = {}
                for col in range(2, 8):  # B through G
                    col_letter = sheet.cell(row=1, column=col).column_letter
                    if col_letter in sheet.column_dimensions:
                        original_widths[col] = sheet.column_dimensions[col_letter].width

                # Remove the first row while preserving the yellow header row
                sheet.delete_rows(1, 1)

                # Set the height of the new header row (row 1) to match the original
                sheet.row_dimensions[1].height = original_height

                # Delete all columns after G (including H)
                max_col = sheet.max_column
                if max_col > 8:  # If there are columns after G
                    sheet.delete_cols(9, max_col - 8)

                # Delete column A (first column)
                sheet.delete_cols(1, 1)

                # Restore original column widths (now shifted one column left due to deletion of column A)
                for original_col, width in original_widths.items():
                    new_col = original_col - 1  # Shift column index left by 1
                    col_letter = sheet.cell(
                        row=1, column=new_col).column_letter
                    sheet.column_dimensions[col_letter].width = width

                # Combine all DataFrames
                self.logger.debug("Combining NAV DataFrames...")
                nav_df = pd.concat(
                    [df for _, df in nav_dfs], ignore_index=True)
                self.logger.info(f"Processing {len(nav_df)} NAV entries")
                self.logger.info(
                    f"Unique ISINs in input data: {len(nav_df['ISIN'].unique())}")

                # Get the most recent date from the input files
                nav_date = nav_df['Valuation Period-End Date'].max()
                formatted_date = nav_date.strftime('%Y.%m.%d')

                # Get series information from database
                with self.db_service.SessionMaker() as session:
                    # Get all unique ISINs from the NAV data
                    isins = nav_df['ISIN'].unique()
                    self.logger.info(
                        f"Querying database for {len(isins)} ISINs")

                    # Query series information for these ISINs
                    series_info = session.query(Series).filter(
                        Series.isin.in_(isins)).all()
                    self.logger.info(
                        f"Found {len(series_info)} matching series in database")

                    # Create a dictionary for quick lookup
                    series_dict = {s.isin: s for s in series_info}

                # Update NAV values for each series
                rows_updated = 0
                # Start from row 2 (right after the yellow header)
                current_row = 2

                for _, row in nav_df.iterrows():
                    isin = row['ISIN']
                    series = series_dict.get(isin)

                    if series:
                        # Fill in the row data in the correct order:
                        # 1. Security name (using series name)
                        sheet.cell(row=current_row, column=1,
                                   value=str(series.series_name).strip())
                        # 2. ISIN
                        sheet.cell(row=current_row, column=2, value=str(isin))
                        # 3. Valuation Date
                        sheet.cell(row=current_row, column=3,
                                   value=row['Valuation Period-End Date'])
                        # 4. Currency
                        sheet.cell(row=current_row, column=4, value=str(
                            series.currency) if series.currency else "USD")
                        # 5. Nav Price
                        sheet.cell(row=current_row, column=5,
                                   value=float(row['NAV']))
                        # 6. Nav Type (always Structured Products)
                        sheet.cell(row=current_row, column=6,
                                   value="Structured Products")
                        # 7. Valuation Frequency
                        sheet.cell(row=current_row, column=7, value=str(
                            series.nav_frequency.value) if series.nav_frequency else "Daily")
                        current_row += 1
                        rows_updated += 1
                    else:
                        self.logger.warning(
                            f"No series information found for ISIN {isin}")

                self.logger.info(
                    f"Updated {rows_updated} rows in the template")

                # Use date-based filename like Morningstar template
                output_path = self.output_dir / \
                    f'LAM_SFI_Price - {formatted_date}.xlsx'

                # Save workbook
                wb.save(str(output_path))
                wb.close()

                self.logger.info(
                    f"Successfully updated SIX template and saved to {output_path}")
                return output_path

            except Exception as e:
                self.logger.error(f"Error updating SIX template: {str(e)}")
                raise
            finally:
                try:
                    os.unlink(tmp_file.name)
                    self.logger.debug("Cleaned up temporary file")
                except:
                    self.logger.warning("Failed to clean up temporary file")
                    pass

    def process_navs(self, date_str: str, send_email: bool = False, to_emails: List[str] = None,
                     isin_filter: Union[str, List[str], None] = None, distribution_type: str = 'morningstar',
                     template_types: List[str] = ['morningstar', 'six']):
        """Process NAV files and update templates.

        Args:
            date_str (str): Date string in format MMDDYYYY
            send_email (bool): Whether to send email report
            to_emails (List[str]): List of email recipients
            isin_filter (Union[str, List[str], None]): ISIN filter specification
            distribution_type (str): Type of distribution
            template_types (List[str]): List of templates to update ('morningstar', 'six')
        """
        try:
            # Clean up directories
            self.template_processor._cleanup_output_directory()

            # Clean up emitter directories
            if self.ftp_service:
                for emitter in self.ftp_service.config.keys():
                    self.ftp_service.cleanup_emitter_directory(
                        emitter, self.input_dir)

            # Process ISIN filters
            target_isins = self._get_target_isins(isin_filter)
            if target_isins:
                self.logger.info(
                    f"Filtering for ISINs: {len(target_isins)} ISINs selected")

            # Read exclude ISINs
            exclude_isins = self._read_exclude_isins()

            # Process NAV files
            nav_dfs = self._process_nav_files(
                date_str, target_isins, exclude_isins)
            if not nav_dfs:
                raise ValueError("No NAV files could be processed")

            output_paths = []

            # Update templates based on specified types
            for template_type in template_types:
                try:
                    if template_type.lower() == 'morningstar':
                        output_path = self.template_processor.update_morningstar_template(
                            nav_dfs, date_str)
                        output_paths.append(output_path)
                    elif template_type.lower() == 'six':
                        # Get series information for SIX template
                        all_isins = set()
                        for _, df in nav_dfs:
                            all_isins.update(df['ISIN'].unique())
                        series_info = self._get_series_info(all_isins)

                        output_path = self.template_processor.update_six_template(
                            nav_dfs, date_str, series_info)
                        output_paths.append(output_path)
                    else:
                        self.logger.warning(
                            f"Unknown template type: {template_type}")
                except Exception as e:
                    self.logger.error(
                        f"Error updating {template_type} template: {str(e)}")
                    raise

            # Handle email sending if requested
            if send_email and to_emails:
                if isinstance(to_emails, str):
                    to_emails = [to_emails]

                # Send separate emails for each template type
                for template_type in template_types:
                    # Find the corresponding output path for this template type
                    template_path = None
                    for path in output_paths:
                        if (template_type.lower() == 'six' and 'LAM_SFI_Price' in path.name) or \
                           (template_type.lower() == 'morningstar' and 'Flexfunds ETPs' in path.name):
                            template_path = path
                            break

                    if template_path:
                        self._send_email_report(
                            output_path=template_path,
                            to_emails=to_emails,
                            distribution_type=template_type.lower(),
                            nav_dfs=nav_dfs
                        )

            # Upload to Drive if configured
            if self.drive_service:
                uploads_count = 0
                for output_path in output_paths:
                    # Upload to appropriate folder based on template type
                    for template_type in template_types:
                        if template_type.lower() == 'morningstar' and self.drive_config.get('morningstar_output_folder_id'):
                            if 'Flexfunds ETPs - NAVs' in output_path.name:
                                uploads_count += 1
                                self.drive_service.upload_file(
                                    output_path,
                                    self.drive_config['morningstar_output_folder_id']
                                )
                        elif template_type.lower() == 'six' and self.drive_config.get('six_output_folder_id'):
                            if 'LAM_SFI_Price' in output_path.name:
                                uploads_count += 1
                                self.drive_service.upload_file(
                                    output_path,
                                    self.drive_config['six_output_folder_id']
                                )

                # Single consolidated log message
                if uploads_count > 0:
                    self.logger.info(
                        f"Uploaded {uploads_count} template(s) to Google Drive")

            # Save to database
            self._save_to_database(nav_dfs, distribution_type)

            # Clean up
            self.cleanup()

        except Exception as e:
            self.logger.error(f"Error in NAV processing: {str(e)}")
            self.cleanup()
            raise

    def _read_exclude_isins(self) -> Set[str]:
        """Read and return set of ISINs to exclude."""
        exclude_isins = set()
        exclude_isins_path = self.template_dir / "Exclude ISINs.csv"
        if exclude_isins_path.exists():
            exclude_isins = set(pd.read_csv(
                exclude_isins_path, header=None)[0].str.strip())
            self.logger.info(f"Loaded {len(exclude_isins)} ISINs to exclude")
        return exclude_isins

    def _get_target_isins(self, isin_filter: Union[str, List[str], None] = None) -> Optional[Set[str]]:
        """Helper method to process ISIN filters

        Args:
            isin_filter: Can be one of:
                - A predefined frequency ("daily", "weekly", "monthly", "quarterly")
                - A list of predefined frequencies (e.g., ["daily", "weekly"])
                - A specific ISIN or list of ISINs
                - None to process all ISINs

        Returns:
            Optional[Set[str]]: Set of target ISINs or None if no filter applied
        """
        if not isin_filter:
            return None

        target_isins = set()

        if isinstance(isin_filter, str):
            # Check if it's a frequency or a specific ISIN
            if isin_filter.upper() in ['DAILY', 'WEEKLY', 'MONTHLY', 'QUARTERLY']:
                target_isins.update(
                    self._get_isins_by_frequency(isin_filter.upper()))
            else:
                target_isins.add(isin_filter)
        else:
            # List of frequencies or ISINs
            for item in isin_filter:
                if item.upper() in ['DAILY', 'WEEKLY', 'MONTHLY', 'QUARTERLY']:
                    target_isins.update(
                        self._get_isins_by_frequency(item.upper()))
                else:
                    target_isins.add(item)

        return target_isins if target_isins else None

    def _get_isins_by_frequency(self, frequency: str) -> List[str]:
        """Get ISINs for a specific NAV frequency."""
        # Convert frequency to uppercase for consistent comparison
        frequency = frequency.upper()
        with self.db_service.SessionMaker() as session:
            return [r[0] for r in session.query(Series.isin)
                    .filter(func.upper(Series.nav_frequency) == frequency)
                    .filter(Series.status == SeriesStatus.ACTIVE)
                    .all()]

    def import_historic_data(self, excel_path: str):
        """
        Import historic NAV data from Excel file

        Args:
            excel_path (str): Path to the Excel file containing historic NAV data
        """
        try:
            results = self.db_service.import_historic_data(excel_path)

            # Calculate total counts across all sheets
            total_added = sum(
                result.added_count for result in results.values())
            total_duplicates = sum(
                result.duplicates_count for result in results.values())

            self.logger.info(
                f"Successfully imported historic NAV data: {total_added} new entries added, "
                f"{total_duplicates} duplicates skipped"
            )
            return total_added, total_duplicates
        except Exception as e:
            self.logger.error(f"Error importing historic NAV data: {str(e)}")
            raise

    def get_nav_history(self, isin: str, start_date: Optional[datetime] = None,
                        end_date: Optional[datetime] = None):
        """
        Get NAV history for a specific ISIN within a date range

        Args:
            isin (str): The ISIN to query
            start_date (Optional[datetime]): Start date for history query
            end_date (Optional[datetime]): End date for history query

        Returns:
            List[NAVEntry]: List of NAV entries matching the query
        """
        try:
            nav_entries = self.db_service.get_nav_history(
                isin, start_date, end_date)
            self.logger.info(
                f"Retrieved {len(nav_entries)} NAV entries for ISIN {isin}")
            return nav_entries
        except Exception as e:
            self.logger.error(f"Error retrieving NAV history: {str(e)}")
            raise

    def set_debug_logging(self, enable=True):
        """Enable or disable detailed debug logging."""
        if enable:
            # Set verbose logging
            self.logger.setLevel(logging.DEBUG)
            logging.getLogger('googleapiclient').setLevel(logging.DEBUG)
            logging.getLogger('google_auth_httplib2').setLevel(logging.DEBUG)
            logging.getLogger('ftplib').setLevel(logging.DEBUG)
        else:
            # Restore normal logging
            self.logger.setLevel(logging.INFO)
            logging.getLogger('googleapiclient').setLevel(logging.WARNING)
            logging.getLogger('google_auth_httplib2').setLevel(logging.WARNING)
            logging.getLogger('ftplib').setLevel(logging.WARNING)


def main():
    # Example usage for remote mode with multiple FTP configs
    ftp_configs = {
        "ETPCAP2": {
            "host": "",
            "user": "nav_auto",
            "password": "hola",
            "directory": "/1"
        },
        "HFMX": {
            "host": "teo.superhosting.bg",
            "user": "data@hfmxdacseries.com",
            "password": "BF0*5bZIRZK^",
            "directory": "/"
        },
        "IACAP": {
            "host": "omar.superhosting.bg",
            "user": "data@iacapitalplc.com",
            "password": "BF0*5bZIRZK^",
            "directory": "/"
        },
        "CIX": {
            "host": "mini.superhosting.bg",
            "user": "data@cixdac.com",
            "password": "BF0*5bZIRZK^",
            "directory": "/"
        },
        "DCX": {
            "host": "mini.superhosting.bg",
            "user": "data@dcxpd.com",
            "password": "9RF#c[tCq}rT",
            "directory": "/"
        }
    }

    # Initialize processor with database support
    processor = NAVProcessor(
        mode="remote",
        ftp_configs=ftp_configs,
        db_connection_string='sqlite:///nav_data.db'
    )

    # Process NAVs and save to database
    date_str = datetime.now().strftime('%m%d%Y')
    processor.process_navs(date_str)

    # Example: Query NAV history for a specific ISIN
    isin = "XS2728487260"  # Example ISIN from daily set
    start_date = datetime.now().replace(day=1)  # First day of current month
    nav_history = processor.get_nav_history(isin, start_date)
    print(f"\nNAV history for {isin} since {start_date.date()}:")
    for entry in nav_history:
        print(f"Date: {entry.nav_date}, Value: {entry.nav_value}")


if __name__ == "__main__":
    main()
